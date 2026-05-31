# -*- coding: utf-8 -*-
"""
시계열 통합 Excel의 무결성 검증 — 결정적 영역.

검사 항목:
  1. BS 등식: 자산총계 = 부채총계 + 자본총계 (별도/연결 각 연도)
  2. IS 등식: 매출액 = 매출원가 + 매출총이익
  3. 핵심 계정 누락: 매출액·자산총계·자본총계가 어느 연도에 비어 있는지

사용:
  py validate.py output/2026-05/무신사_시계열_2026-05-17.xlsx
  py validate.py output/2026-05/무신사_시계열_2026-05-17.xlsx --log _workspace/무신사/logs

종료 코드:
  0  통과
  1  검증 실패 (등식 어긋남 또는 누락 있음)
  2  파일/포맷 에러
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field, asdict
from datetime import date
from pathlib import Path

if __name__ == "__main__":
    for _s in (sys.stdout, sys.stderr):
        try:
            if hasattr(_s, "reconfigure"):
                _s.reconfigure(encoding="utf-8")
        except Exception:
            pass

import openpyxl
from openpyxl.styles import PatternFill, Font


TOL = 100   # 100원 미만 차이는 반올림 노이즈로 간주

# 검증_Report 시트 셀 색상/라벨 (severity → 표현)
SEV_FILL = {
    "ok":   PatternFill("solid", fgColor="C6EFCE"),   # 초록
    "warn": PatternFill("solid", fgColor="FFEB9C"),   # 노랑
    "fail": PatternFill("solid", fgColor="FFC7CE"),   # 빨강
}
SEV_LABEL = {"ok": "✓ 일치", "warn": "△ 확인필요", "fail": "✗ 불일치"}


@dataclass
class CheckResult:
    sheet: str
    check: str          # "BS등식" / "IS등식" / "누락"
    year: int | str | None
    detail: str
    severity: str       # "ok" / "warn" / "fail"


@dataclass
class Report:
    file: str
    checked_at: str
    summary: dict = field(default_factory=dict)
    issues: list[CheckResult] = field(default_factory=list)

    def to_dict(self):
        d = asdict(self)
        d["issues"] = [asdict(i) for i in self.issues]
        return d


def _read_pivot(ws) -> tuple[list, dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}
    header = rows[0]
    data = {r[0]: r[1:] for r in rows[1:] if r and r[0]}
    return header, data


def check_bs(ws, sheet_name: str) -> list[CheckResult]:
    header, data = _read_pivot(ws)
    out: list[CheckResult] = []
    for i, year in enumerate(header[1:]):
        a = data.get("자산총계", [None])[i] if data.get("자산총계") else None
        l = data.get("부채총계", [None])[i] if data.get("부채총계") else None
        e = data.get("자본총계", [None])[i] if data.get("자본총계") else None
        if a is None or l is None or e is None:
            out.append(CheckResult(sheet_name, "BS등식", year,
                                    f"자산={a} 부채={l} 자본={e} (일부 누락)", "warn"))
            continue
        diff = a - l - e
        if abs(diff) < TOL:
            out.append(CheckResult(sheet_name, "BS등식", year, f"diff={diff}", "ok"))
        else:
            out.append(CheckResult(sheet_name, "BS등식", year,
                                    f"diff={diff:,} (자산 {a:,} - 부채 {l:,} - 자본 {e:,})", "fail"))
    return out


def check_is(ws, sheet_name: str) -> list[CheckResult]:
    header, data = _read_pivot(ws)
    out: list[CheckResult] = []
    for i, year in enumerate(header[1:]):
        s = data.get("매출액", [None])[i] if data.get("매출액") else None
        c = data.get("매출원가", [None])[i] if data.get("매출원가") else None
        g = data.get("매출총이익", [None])[i] if data.get("매출총이익") else None
        if s is None or c is None or g is None:
            # 서비스업처럼 매출원가가 없는 양식도 있음 → warn까지만
            out.append(CheckResult(sheet_name, "IS등식", year,
                                    f"매출={s} 원가={c} GP={g} (일부 누락)", "warn"))
            continue
        diff = s - c - g
        if abs(diff) < TOL:
            out.append(CheckResult(sheet_name, "IS등식", year, f"diff={diff}", "ok"))
        else:
            out.append(CheckResult(sheet_name, "IS등식", year,
                                    f"diff={diff:,} (매출 {s:,} - 원가 {c:,} - GP {g:,})", "fail"))
    return out


def check_missing_core(ws, sheet_name: str, core_accounts: list[str]) -> list[CheckResult]:
    """핵심 계정이 어느 연도에 비어 있는지 표시."""
    header, data = _read_pivot(ws)
    out: list[CheckResult] = []
    for acc in core_accounts:
        vals = data.get(acc)
        if vals is None:
            out.append(CheckResult(sheet_name, "누락", None, f"계정 자체가 시트에 없음: {acc}", "warn"))
            continue
        missing_years = [header[1 + i] for i, v in enumerate(vals) if v is None]
        if missing_years:
            out.append(CheckResult(sheet_name, "누락", None,
                                    f"{acc} 누락 연도: {missing_years}", "warn"))
    return out


def validate(xlsx_path: Path) -> Report:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rep = Report(file=str(xlsx_path), checked_at=date.today().isoformat())

    bs_sheets = [s for s in ("별도_재무상태표", "연결_재무상태표") if s in wb.sheetnames]
    is_sheets = [s for s in ("별도_포괄손익계산서", "연결_포괄손익계산서") if s in wb.sheetnames]

    for sn in bs_sheets:
        rep.issues.extend(check_bs(wb[sn], sn))
        rep.issues.extend(check_missing_core(wb[sn], sn, ["자산총계", "자본총계"]))
    for sn in is_sheets:
        rep.issues.extend(check_is(wb[sn], sn))
        rep.issues.extend(check_missing_core(wb[sn], sn, ["매출액", "영업이익", "당기순이익"]))

    rep.summary = {
        "ok": sum(1 for i in rep.issues if i.severity == "ok"),
        "warn": sum(1 for i in rep.issues if i.severity == "warn"),
        "fail": sum(1 for i in rep.issues if i.severity == "fail"),
    }
    return rep


def render_report_sheet(rep: Report, xlsx_path: Path, sheet_name: str = "검증_Report") -> Path:
    """검증 결과(Report)를 같은 xlsx 파일에 "검증_Report" 시트로 주입한다.

    - 이미 시트가 있으면 교체 (재실행 안전).
    - 시트는 맨 앞에 배치 → 파일을 열면 정합성이 가장 먼저 보임.
    - 결과 셀은 severity별로 색칠(초록/노랑/빨강).

    결정성 분리: 이 함수는 validate()가 산출한 Report만 렌더링한다. 숫자를 새로
    계산하지 않는다.
    """
    xlsx_path = Path(xlsx_path)
    # 값만 들어있는 pandas 산출물이라 data_only=False로 열어도 값이 보존된다.
    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)

    # 1) 요약 헤더
    ws["A1"] = "검증 요약"
    ws["A1"].font = Font(bold=True, size=12)
    ws["B1"] = (f"통과 {rep.summary.get('ok', 0)} · "
                f"경고 {rep.summary.get('warn', 0)} · "
                f"실패 {rep.summary.get('fail', 0)}")
    ws["B1"].fill = SEV_FILL["fail"] if rep.summary.get("fail", 0) else SEV_FILL["ok"]
    ws["E1"] = f"검증일 {rep.checked_at}"

    # 2) 표 헤더
    hdr = ["재무제표", "연도", "검사", "결과", "상세"]
    head_row = 3
    head_fill = PatternFill("solid", fgColor="D9D9D9")
    for j, h in enumerate(hdr):
        c = ws.cell(row=head_row, column=1 + j, value=h)
        c.font = Font(bold=True)
        c.fill = head_fill

    # 3) 이슈 행: 실패 → 경고 → 통과 순으로 정렬해 중요한 게 위로
    order = {"fail": 0, "warn": 1, "ok": 2}
    issues = sorted(
        rep.issues,
        key=lambda i: (order.get(i.severity, 9), str(i.sheet), str(i.year)),
    )
    r = head_row + 1
    for it in issues:
        ws.cell(row=r, column=1, value=it.sheet)
        ws.cell(row=r, column=2, value=it.year)
        ws.cell(row=r, column=3, value=it.check)
        rc = ws.cell(row=r, column=4, value=SEV_LABEL.get(it.severity, it.severity))
        rc.fill = SEV_FILL.get(it.severity, PatternFill())
        ws.cell(row=r, column=5, value=it.detail)
        r += 1

    for col, width in zip("ABCDE", (22, 8, 10, 12, 64)):
        ws.column_dimensions[col].width = width

    wb.save(xlsx_path)
    return xlsx_path


def validate_and_embed(xlsx_path: Path, log_dir: Path | None = None) -> Report:
    """검증 → 검증_Report 시트 주입 → (선택) JSON 로그 저장을 한 번에.

    산출 경로(build_timeseries/extractor)가 호출하는 단일 진입점.
    """
    xlsx_path = Path(xlsx_path)
    rep = validate(xlsx_path)
    render_report_sheet(rep, xlsx_path)
    if log_dir is not None:
        write_log(rep, Path(log_dir))
    return rep


def print_report(rep: Report):
    print(f"[검증] {rep.file}")
    print(f"        통과 {rep.summary['ok']} / 경고 {rep.summary['warn']} / 실패 {rep.summary['fail']}")
    by_sheet: dict[str, list[CheckResult]] = {}
    for i in rep.issues:
        by_sheet.setdefault(i.sheet, []).append(i)
    for sheet, items in by_sheet.items():
        print(f"  ── {sheet} ──")
        for it in items:
            sym = {"ok": "✓", "warn": "△", "fail": "✗"}[it.severity]
            year = f"{it.year} " if it.year is not None else ""
            print(f"    {sym} [{it.check}] {year}{it.detail}")


def write_log(rep: Report, log_dir: Path) -> Path:
    log_dir.mkdir(parents=True, exist_ok=True)
    fname = f"validation_{rep.checked_at}.log"
    p = log_dir / fname
    with p.open("w", encoding="utf-8") as f:
        json.dump(rep.to_dict(), f, ensure_ascii=False, indent=2)
    return p


def main(argv: list[str]):
    p = argparse.ArgumentParser(description="시계열 Excel 무결성 검증")
    p.add_argument("xlsx", help="검증할 시계열 Excel 경로")
    p.add_argument("--log", help="로그 디렉토리 (지정 시 JSON 로그 저장)")
    p.add_argument("--embed", action="store_true",
                   help="검증 결과를 같은 xlsx에 '검증_Report' 시트로 주입")
    args = p.parse_args(argv[1:])

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"[!] 파일 없음: {xlsx_path}", file=sys.stderr)
        sys.exit(2)

    try:
        rep = validate(xlsx_path)
    except Exception as e:
        print(f"[!] 검증 중 에러: {e}", file=sys.stderr)
        sys.exit(2)

    print_report(rep)

    if args.embed:
        render_report_sheet(rep, xlsx_path)
        print(f"\n검증_Report 시트 주입: {xlsx_path}")

    if args.log:
        log_path = write_log(rep, Path(args.log))
        print(f"로그 저장: {log_path}")

    sys.exit(1 if rep.summary["fail"] > 0 else 0)


if __name__ == "__main__":
    main(sys.argv)
