# -*- coding: utf-8 -*-
"""
DART 감사보고서 → 재무제표 시계열 Excel: 원-샷 자동화 CLI.

회사명 한 줄 입력하면 DART에서 모든 감사보고서/연결감사보고서 PDF를 다운받아
→ 재무제표 4종 Excel로 변환 → 시계열 통합 Excel까지 생성.

대상: 사업보고서 미제출 비상장 외부감사 회사, 상장사의 비상장 시기 자료.

사용법:
  $env:DART_API_KEY="..."
  py extractor.py "카카오스타일"
  py extractor.py --search "쿠팡"             # 검색만
  py extractor.py "카카오스타일" --from 2018 --to 2024
  py extractor.py "카카오스타일" --only separate     # 별도만
  py extractor.py "카카오스타일" --only consolidated # 연결만
  py extractor.py "카카오스타일" --download-only      # PDF만 받고 변환 안 함
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from datetime import date
from pathlib import Path

# 콘솔 한글 깨짐 방지: PYTHONIOENCODING=utf-8 환경변수가 가장 호환성 높다.
# (PowerShell에서 py 실행 시 sys.stdout 재포장은 'closed file' 에러를 종종 일으킴.)
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

import dart_client
import pdf_to_excel
import consolidate


HERE = Path(__file__).resolve().parent
DATA_DIR = HERE / "data"
CORPCODE_XML = DATA_DIR / "CORPCODE.xml"

# 하네스 모드 경로 — CLAUDE.md의 I/O 규약과 동기화
HARNESS_INPUT_RAW = HERE / "_input" / "raw"
HARNESS_WORKSPACE = HERE / "_workspace"
HARNESS_OUTPUT = HERE / "output"


# ---------- 사용자 선택 ----------

def _choose_company(cands: list[dict], keyword: str) -> dict | None:
    """후보가 여러 개일 때 대화형 선택. 1개면 자동 선택."""
    if not cands:
        print(f"[!] '{keyword}' 검색 결과 없음.")
        return None
    if len(cands) == 1:
        return cands[0]
    # 정확 일치가 있고 그게 유일하면 자동 선택
    exact = [c for c in cands if c["corp_name"] == keyword]
    if len(exact) == 1:
        print(f"  → 정확 일치 자동 선택: {exact[0]['corp_name']}")
        return exact[0]

    print(f"\n'{keyword}' 검색 결과 {len(cands)}건:")
    for i, c in enumerate(cands, 1):
        listed = c["stock_code"].strip() or "비상장"
        print(f"  [{i}] {c['corp_name']:30s}  corp_code={c['corp_code']}  {listed}")
    while True:
        ans = input("선택 (번호 또는 q): ").strip()
        if ans.lower() == "q":
            return None
        if ans.isdigit() and 1 <= int(ans) <= len(cands):
            return cands[int(ans) - 1]
        print("  잘못된 입력. 다시 시도.")


# ---------- 파이프라인 ----------

def _safe_filename(s: str) -> str:
    bad = '<>:"/\\|?*'
    return "".join("_" if c in bad else c for c in s).strip()


def _pdf_filename(d: "dart_client.Disclosure") -> str:
    # 예: "감사보고서_2023.12_20240409000213.pdf"
    period = d.report_nm.split("(")[-1].rstrip(")") if "(" in d.report_nm else ""
    kind = "연결감사보고서" if d.kind == "연결" else "감사보고서"
    return _safe_filename(f"{kind}_{period}_{d.rcept_no}.pdf")


def run_pipeline(
    keyword: str,
    *,
    api_key: str,
    year_from: int = 2010,
    year_to: int | None = None,
    only: str | None = None,           # None | "separate" | "consolidated"
    download_only: bool = False,
    harness: bool = False,
) -> Path | None:
    if year_to is None:
        year_to = date.today().year
    bgn_de = f"{year_from}0101"
    end_de = f"{year_to}1231"

    # 1) corpCode 캐시
    print(f"[1/6] corpCode 캐시 확인…")
    dart_client.ensure_corpcode(api_key, CORPCODE_XML)

    # 2) 회사 검색
    print(f"[2/6] '{keyword}' 검색…")
    cands = dart_client.find_companies(keyword, CORPCODE_XML)
    target = _choose_company(cands, keyword)
    if not target:
        return None
    print(f"      선택: {target['corp_name']} (corp_code={target['corp_code']})")

    company_safe = _safe_filename(target["corp_name"])
    today_iso = date.today().isoformat()              # YYYY-MM-DD
    yyyy_mm = today_iso[:7]                            # YYYY-MM

    if harness:
        # 하네스 경로: _input/raw/ (read-only PDF), _workspace/ (중간), output/YYYY-MM/ (최종)
        pdf_dir = HARNESS_INPUT_RAW / company_safe
        excel_dir = HARNESS_WORKSPACE / company_safe / "excels"
        out_xlsx = HARNESS_OUTPUT / yyyy_mm / f"{company_safe}_시계열_{today_iso}.xlsx"
        print(f"      [harness 모드] raw={pdf_dir.relative_to(HERE)}  "
              f"workspace={excel_dir.relative_to(HERE)}  out={out_xlsx.relative_to(HERE)}")
    else:
        # 레거시 경로: data/<회사>/...
        work_dir = DATA_DIR / company_safe
        pdf_dir = work_dir / "pdfs"
        excel_dir = work_dir / "excels"
        out_xlsx = work_dir / f"{company_safe}_시계열.xlsx"

    # 3) 공시 목록
    print(f"[3/6] 감사보고서 공시 조회 ({bgn_de}~{end_de})…")
    disclosures = dart_client.list_audit_disclosures(
        api_key, target["corp_code"], bgn_de=bgn_de, end_de=end_de
    )
    # 정정공시 dedupe → 사업연도별로 최신 한 건씩
    disclosures = dart_client.dedupe_latest(disclosures)
    # 별도/연결 필터
    if only == "separate":
        disclosures = [d for d in disclosures if d.kind == "별도"]
    elif only == "consolidated":
        disclosures = [d for d in disclosures if d.kind == "연결"]
    print(f"      대상 보고서: 별도 {sum(1 for d in disclosures if d.kind=='별도')}건, "
          f"연결 {sum(1 for d in disclosures if d.kind=='연결')}건")
    if not disclosures:
        print("[!] 해당 기간에 감사보고서 없음.")
        return None

    # 4) PDF 다운로드
    print(f"[4/6] PDF 다운로드 → {pdf_dir}")
    import requests
    sess = requests.Session()
    for d in disclosures:
        dest = pdf_dir / _pdf_filename(d)
        if dest.exists() and dest.stat().st_size > 1024:
            print(f"      [skip] {dest.name} (이미 있음)")
            continue
        try:
            dart_client.download_pdf(d.rcept_no, dest, session=sess)
            print(f"      ✓ {dest.name} ({dest.stat().st_size//1024} KB)")
        except Exception as e:
            print(f"      [에러] {d.rcept_no} ({d.report_nm}): {e}")
        time.sleep(0.3)

    if download_only:
        print("\n다운로드만 요청됨. 변환 스킵.")
        return pdf_dir

    # 5) PDF → Excel 변환
    print(f"[5/6] PDF → Excel 변환 → {excel_dir}")
    excel_dir.mkdir(parents=True, exist_ok=True)
    pdfs = sorted(pdf_dir.glob("*.pdf"))
    for pdf in pdfs:
        out = excel_dir / f"{pdf.stem}.xlsx"
        if out.exists():
            print(f"      [skip] {out.name}")
            continue
        try:
            fs = pdf_to_excel.extract_financials(pdf)
            pdf_to_excel.save_excel(fs, out)
            found = [n for n in ("재무상태표", "포괄손익계산서", "자본변동표", "현금흐름표") if fs.get(n) is not None]
            print(f"      ✓ {out.name}  ({len(found)}/4 시트)")
        except Exception as e:
            print(f"      [에러] {pdf.name}: {e}")

    # 6) 시계열 통합
    print(f"[6/6] 시계열 통합 → {out_xlsx.name}")
    consolidate.run(excel_dir, out_xlsx, default_company=target["corp_name"])

    # 회계등식 자동 검증
    _verify_accounting(out_xlsx)

    return out_xlsx


def _verify_accounting(xlsx_path: Path):
    """자산=부채+자본, 매출=원가+매출총이익 자동 검증."""
    try:
        import openpyxl
    except ImportError:
        return
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"\n[검증] {xlsx_path.name}")

    for sn in ("별도_재무상태표", "연결_재무상태표"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data = {r[0]: r[1:] for r in rows[1:]}
        ok = 0; warn = []
        for i, year in enumerate(header[1:]):
            a = data.get("자산총계", [None])[i]
            l = data.get("부채총계", [None])[i]
            e = data.get("자본총계", [None])[i]
            if a and l and e and abs(a - l - e) < 100:
                ok += 1
            elif a and l and e:
                warn.append(f"{year}: diff={a-l-e:,}")
        print(f"   {sn}: BS등식 {ok}/{len(header)-1} ✓  {('|'.join(warn)) if warn else ''}")

    for sn in ("별도_포괄손익계산서", "연결_포괄손익계산서"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        data = {r[0]: r[1:] for r in rows[1:]}
        ok = 0; missing = []
        for i, year in enumerate(header[1:]):
            s = data.get("매출액", [None])[i]
            c = data.get("매출원가", [None])[i]
            g = data.get("매출총이익", [None])[i]
            if s and c and g and abs(s - c - g) < 100:
                ok += 1
            elif not (s and c and g):
                missing.append(str(year))
        msg = f"   {sn}: IS등식 {ok}/{len(header)-1} ✓"
        if missing:
            msg += f"  (일부누락: {', '.join(missing)})"
        print(msg)


# ---------- CLI ----------

def main(argv: list[str]):
    p = argparse.ArgumentParser(
        prog="extractor.py",
        description="DART 감사보고서 → 재무제표 시계열 Excel 원-샷 도구",
    )
    p.add_argument("company", nargs="?", help="회사명 (또는 키워드)")
    p.add_argument("--search", metavar="키워드", help="회사 검색만 수행하고 종료")
    p.add_argument("--key", help="DART API 키 (없으면 환경변수 DART_API_KEY)")
    p.add_argument("--from", dest="year_from", type=int, default=2010, help="시작 연도 (공시 접수일 기준)")
    p.add_argument("--to",   dest="year_to",   type=int, default=None, help="종료 연도 (공시 접수일 기준)")
    p.add_argument("--only", choices=["separate", "consolidated"], help="별도 또는 연결 한쪽만")
    p.add_argument("--download-only", action="store_true", help="PDF만 받고 Excel 변환·통합 스킵")
    p.add_argument("--harness", action="store_true",
                   help="하네스 경로 사용 (_input/raw/, _workspace/, output/YYYY-MM/). 기본은 data/ 레거시 경로.")
    args = p.parse_args(argv[1:])

    api_key = (args.key or os.environ.get("DART_API_KEY", "")).strip()
    if not api_key:
        print("[!] DART API 키가 필요합니다. --key 또는 환경변수 DART_API_KEY 로 전달하세요.")
        print("    https://opendart.fss.or.kr 에서 무료 발급.")
        sys.exit(1)

    if args.search:
        dart_client.ensure_corpcode(api_key, CORPCODE_XML)
        cands = dart_client.find_companies(args.search, CORPCODE_XML)
        print(f"'{args.search}' 검색 결과 {len(cands)}건:")
        for c in cands:
            listed = c["stock_code"].strip() or "비상장"
            print(f"  {c['corp_name']:30s}  corp_code={c['corp_code']}  {listed}  modify={c['modify_date']}")
        return

    if not args.company:
        p.print_help()
        sys.exit(1)

    out = run_pipeline(
        args.company,
        api_key=api_key,
        year_from=args.year_from,
        year_to=args.year_to,
        only=args.only,
        download_only=args.download_only,
        harness=args.harness,
    )
    if out:
        print(f"\n✓ 완료: {out}")


if __name__ == "__main__":
    main(sys.argv)
